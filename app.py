from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import sqlite3
import os
from datetime import datetime
import uuid
import json
import numpy as np  
import pandas as pd
from io import BytesIO
from flask import send_file
from datetime import datetime, timedelta
import tempfile


# Face recognition imports (optional)
try:
    import cv2
    import face_recognition
    FACE_RECOGNITION_AVAILABLE = True
except ImportError:
    FACE_RECOGNITION_AVAILABLE = False
    print("Warning: Face recognition libraries not installed. Install with:")
    print("pip install opencv-python face_recognition")

# Import custom modules
from register_web import init_web_registration

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'  # Ganti dengan secret key yang aman
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['FACES_FOLDER'] = 'faces'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Allowed file extensions
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db_connection():
    """Get database connection"""
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn

def login_required(f):
    """Decorator to require login for routes"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def verify_face_for_attendance(image_file, user_id):
    """Verify face for attendance"""
    if not FACE_RECOGNITION_AVAILABLE:
        return True, "Face recognition not available, skipping verification"
    
    try:
        # Get stored face encoding from database
        conn = get_db_connection()
        face_data = conn.execute(
            'SELECT face_encoding FROM face_data WHERE user_id = ? AND active = 1',
            (user_id,)
        ).fetchone()
        conn.close()
        
        if not face_data:
            # No face data stored, allow attendance but warn
            return True, "No face data registered, attendance allowed"
        
        # Load stored encoding
        stored_encoding = np.array(json.loads(face_data[0]))
        
        # Process uploaded image
        image = face_recognition.load_image_file(image_file)
        face_encodings = face_recognition.face_encodings(image)
        
        if not face_encodings:
            return False, "Wajah tidak terdeteksi."
        
        if len(face_encodings) > 1:
            return False, "Terdeteksi lebih dari satu wajah!"
        
        # Compare faces
        matches = face_recognition.compare_faces([stored_encoding], face_encodings[0])
        face_distance = face_recognition.face_distance([stored_encoding], face_encodings[0])
        
        # Threshold for face matching (lower = more strict)
        threshold = 0.4
        
        if matches[0] and face_distance[0] < threshold:
            confidence = (1 - face_distance[0]) * 100
            return True, f"Wajah terverifikasi! Akurasi: {confidence:.1f}%"
        else:
            return False, f"Wajah tidak dikenali."
            
    except Exception as e:
        return False, f"Error verifying face: {str(e)}"

def cleanup_old_attendance_photos(days_to_keep=7):
    """
    Delete attendance photos older than specified days
    Args:
        days_to_keep: Number of days to keep photos (default 7)
    """
    try:
        conn = get_db_connection()
        cutoff_date = (datetime.now() - timedelta(days=days_to_keep)).strftime('%Y-%m-%d')
        
        # Get old attendance records with photos
        old_photos = conn.execute('''
            SELECT photo_path, photo_path_out 
            FROM attendance 
            WHERE date < ?
        ''', (cutoff_date,)).fetchall()
        
        deleted_count = 0
        
        for record in old_photos:
            # Delete check-in photo
            if record['photo_path'] and os.path.exists(record['photo_path']):
                try:
                    os.remove(record['photo_path'])
                    deleted_count += 1
                except Exception as e:
                    print(f"Error deleting {record['photo_path']}: {e}")
            
            # Delete check-out photo
            if record['photo_path_out'] and os.path.exists(record['photo_path_out']):
                try:
                    os.remove(record['photo_path_out'])
                    deleted_count += 1
                except Exception as e:
                    print(f"Error deleting {record['photo_path_out']}: {e}")
        
        # Update database - set photo paths to NULL for old records
        conn.execute('''
            UPDATE attendance 
            SET photo_path = NULL, photo_path_out = NULL 
            WHERE date < ?
        ''', (cutoff_date,))
        
        conn.commit()
        conn.close()
        
        print(f"Cleanup complete: {deleted_count} photos deleted (older than {days_to_keep} days)")
        return deleted_count
        
    except Exception as e:
        print(f"Error during cleanup: {e}")
        return 0

@app.route('/absensi')
@login_required
def absensi():
    """Attendance page"""
    conn = get_db_connection()
    
    # Get today's attendance
    today = datetime.now().strftime("%Y-%m-%d")
    attendance = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? AND date = ?',
        (session['user_id'], today)
    ).fetchone()
    
    # Get allowed coordinates - TAMBAHKAN INI!
    coordinates_rows = conn.execute('SELECT * FROM coordinates WHERE active = 1').fetchall()
    
    # Convert Row objects to dict untuk JSON serialization
    coordinates = [dict(row) for row in coordinates_rows]
    
    # Check if user has face recognition enabled
    face_enabled = conn.execute(
        'SELECT COUNT(*) FROM face_data WHERE user_id = ? AND active = 1',
        (session['user_id'],)
    ).fetchone()[0] > 0
    
    conn.close()
    
    return render_template('absensi.html', 
                         attendance=attendance, 
                         coordinates=coordinates,  # PASS coordinates ke template
                         face_enabled=face_enabled,
                         face_recognition_available=FACE_RECOGNITION_AVAILABLE)

     
import math
def haversine(lat1, lon1, lat2, lon2):
    """
    Hitung jarak antara dua titik koordinat (meter)
    """
    R = 6371000  # radius bumi dalam meter
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)

    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    return R * (2 * math.atan2(math.sqrt(a), math.sqrt(1-a)))


@app.route('/absen_masuk', methods=['POST'])
@login_required
def absen_masuk():
    """Clock in endpoint with mandatory face verification"""
    try:
        latitude = float(request.form.get('latitude', 0))
        longitude = float(request.form.get('longitude', 0))
        
        conn = get_db_connection()
        today = datetime.now().strftime("%Y-%m-%d")
        now = datetime.now().strftime("%H:%M:%S")

        # ✅ CEK WAJIB: User harus sudah setup face recognition
        face_enabled = conn.execute(
            'SELECT COUNT(*) FROM face_data WHERE user_id = ? AND active = 1',
            (session['user_id'],)
        ).fetchone()[0] > 0
        
        if not face_enabled:
            conn.close()
            return jsonify({
                'success': False, 
                'message': 'Anda harus setup Face Recognition terlebih dahulu di menu Profil!',
                'require_face_setup': True
            })

        # Validasi lokasi
        coordinates = conn.execute('SELECT * FROM coordinates WHERE active = 1').fetchall()
        in_area = False
        for coord in coordinates:
            distance = haversine(latitude, longitude, coord['latitude'], coord['longitude'])
            if distance <= coord['radius']:
                in_area = True
                break

        if not in_area:
            conn.close()
            return jsonify({'success': False, 'message': 'Anda berada di luar area absensi!'})

        # Cek sudah absen masuk
        existing = conn.execute(
            'SELECT id FROM attendance WHERE user_id = ? AND date = ?',
            (session['user_id'], today)
        ).fetchone()
        
        if existing:
            conn.close()
            return jsonify({'success': False, 'message': 'Anda sudah absen hari ini!'})

        # ✅ WAJIB: Foto untuk face recognition
        if 'photo' not in request.files:
            conn.close()
            return jsonify({
                'success': False, 
                'message': 'Foto wajib diperlukan untuk verifikasi identitas!'
            })
        
        file = request.files['photo']
        if not file or not allowed_file(file.filename):
            conn.close()
            return jsonify({
                'success': False, 
                'message': 'File foto tidak valid!'
            })
        
        # Save photo
        filename = f"{session['user_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(photo_path)
        
        # ✅ Verifikasi wajah (WAJIB jika face recognition available)
        if FACE_RECOGNITION_AVAILABLE:
            face_verified, face_message = verify_face_for_attendance(photo_path, session['user_id'])
            if not face_verified:
                if os.path.exists(photo_path):
                    os.remove(photo_path)
                conn.close()
                return jsonify({'success': False, 'message': f'{face_message}'})
        else:
            # Jika library tidak tersedia, tetap izinkan tapi log warning
            face_message = "Face recognition library not available"
        
        conn.execute(
            '''INSERT INTO attendance (user_id, date, time_in, latitude, longitude, photo_path)
               VALUES (?, ?, ?, ?, ?, ?)''',
            (session['user_id'], today, now, latitude, longitude, photo_path)
        )

        conn.execute(
            '''INSERT INTO attendance_logs (user_id, action, latitude, longitude, success)
               VALUES (?, ?, ?, ?, ?)''',
            (session['user_id'], 'check_in', latitude, longitude, 1)
        )

        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': f'Absen masuk berhasil! {face_message if FACE_RECOGNITION_AVAILABLE else ""}'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

    
    
@app.route('/absen_keluar', methods=['POST'])
@login_required
def absen_keluar():
    """Clock out endpoint with mandatory face verification"""
    try:
        latitude = float(request.form.get('latitude', 0))
        longitude = float(request.form.get('longitude', 0))

        conn = get_db_connection()
        today = datetime.now().strftime("%Y-%m-%d")
        now = datetime.now().strftime("%H:%M:%S")

        # ✅ CEK WAJIB: User harus sudah setup face recognition
        face_enabled = conn.execute(
            'SELECT COUNT(*) FROM face_data WHERE user_id = ? AND active = 1',
            (session['user_id'],)
        ).fetchone()[0] > 0
        
        if not face_enabled:
            conn.close()
            return jsonify({
                'success': False, 
                'message': 'Anda harus setup Face Recognition terlebih dahulu di menu Profil!',
                'require_face_setup': True
            })

        # Get user info
        user = conn.execute(
            'SELECT u.*, COALESCE(c.name, "Tidak ada kelas") as class_name FROM users u LEFT JOIN classes c ON u.class_id = c.id WHERE u.id = ?',
            (session['user_id'],)
        ).fetchone()

        attendance = conn.execute(
            'SELECT id, time_in, time_out FROM attendance WHERE user_id = ? AND date = ?',
            (session['user_id'], today)
        ).fetchone()

        if not attendance:
            conn.close()
            return jsonify({'success': False, 'message': 'Anda belum absen masuk hari ini!'})

        if attendance['time_out']:
            conn.close()
            return jsonify({'success': False, 'message': 'Anda sudah absen keluar hari ini!'})

        # Validasi lokasi
        coordinates = conn.execute('SELECT * FROM coordinates WHERE active = 1').fetchall()
        in_area = False
        for coord in coordinates:
            distance = haversine(latitude, longitude, coord['latitude'], coord['longitude'])
            if distance <= coord['radius']:
                in_area = True
                break
                
        if not in_area:
            conn.close()
            return jsonify({'success': False, 'message': 'Anda berada di luar area absensi!'})

        # ✅ WAJIB: Foto untuk face recognition
        if 'photo' not in request.files:
            conn.close()
            return jsonify({
                'success': False, 
                'message': 'Foto wajib diperlukan untuk verifikasi identitas!'
            })
        
        file = request.files['photo']
        if not file or not allowed_file(file.filename):
            conn.close()
            return jsonify({
                'success': False, 
                'message': 'File foto tidak valid!'
            })
        
        filename = f"{session['user_id']}_keluar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        photo_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(photo_path)
        
        # ✅ Verifikasi wajah (WAJIB)
        if FACE_RECOGNITION_AVAILABLE:
            face_verified, face_message = verify_face_for_attendance(photo_path, session['user_id'])
            if not face_verified:
                if os.path.exists(photo_path):
                    os.remove(photo_path)
                conn.close()
                return jsonify({'success': False, 'message': f'{face_message}'})
        else:
            face_message = "Face recognition library not available"

        conn.execute(
            'UPDATE attendance SET time_out = ?, latitude_out = ?, longitude_out = ?, photo_path_out = ? WHERE id = ?',
            (now, latitude, longitude, photo_path, attendance['id'])
        )

        conn.execute(
            '''INSERT INTO attendance_logs (user_id, action, latitude, longitude, success)
               VALUES (?, ?, ?, ?, ?)''',
            (session['user_id'], 'check_out', latitude, longitude, 1)
        )

        conn.commit()
        
        # Calculate duration
        time_in = datetime.strptime(attendance['time_in'], '%H:%M:%S')
        time_out = datetime.strptime(now, '%H:%M:%S')
        duration_seconds = (time_out - time_in).total_seconds()
        hours = int(duration_seconds // 3600)
        minutes = int((duration_seconds % 3600) // 60)
        duration_text = f"{hours} jam {minutes} menit"
        
        conn.close()

        return jsonify({
            'success': True, 
            'message': f'Absen keluar berhasil! {face_message if FACE_RECOGNITION_AVAILABLE else ""}',
            'user_name': user['full_name'],
            'class_name': user['class_name'],
            'time_in': attendance['time_in'],
            'time_out': now,
            'duration': duration_text
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    
from datetime import datetime
@app.route('/profil', methods=['GET', 'POST'])
@login_required
def profil():
    """User profile management"""
    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    
    # Check if user has face recognition enabled
    face_data = conn.execute(
        'SELECT * FROM face_data WHERE user_id = ? AND active = 1',
        (session['user_id'],)
    ).fetchone()
    
    if request.method == 'POST':
        full_name = request.form['full_name']
        email = request.form['email']
        password = request.form.get('password', '')  # ✅ UBAH: gunakan .get() untuk optional password

        if password:  # update dengan password
            hashed_pw = generate_password_hash(password)
            conn.execute("""
                UPDATE users SET full_name=?, email=?, password=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (full_name, email, hashed_pw, session['user_id']))
        else:  # update tanpa password
            conn.execute("""
                UPDATE users SET full_name=?, email=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (full_name, email, session['user_id']))
        
        # Update session data
        session['full_name'] = full_name
        
        conn.commit()
        conn.close()
        flash("Profil berhasil diperbarui!", "success")
        return redirect(url_for('profil'))

    # Convert user to dict and parse created_at if it exists
    user_dict = dict(user) if user else {}
    if user_dict.get('created_at'):
        try:
            user_dict['created_at'] = datetime.fromisoformat(user_dict['created_at'].replace('Z', '+00:00'))
        except:
            user_dict['created_at'] = None

    conn.close()
    return render_template('profil.html', 
                         user=user_dict, 
                         face_data=face_data,
                         face_recognition_available=FACE_RECOGNITION_AVAILABLE)
    
    
@app.route('/setup_face', methods=['POST'])
@login_required
def setup_face():
    """Setup face recognition for user"""
    if not FACE_RECOGNITION_AVAILABLE:
        return jsonify({'success': False, 'message': 'Face recognition not available'})
    
    if 'face_image' not in request.files:
        return jsonify({'success': False, 'message': 'No face image provided'})
    
    face_file = request.files['face_image']
    if face_file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not allowed_file(face_file.filename):
        return jsonify({'success': False, 'message': 'Invalid file format'})
    
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        
        # Create user-specific folder
        user_folder = os.path.join(app.config['FACES_FOLDER'], f"{user['full_name']}_{user['id']}")
        if not os.path.exists(user_folder):
            os.makedirs(user_folder)
        
        # Save original image
        filename = secure_filename(f"{user['id']}_face.jpg")
        image_path = os.path.join(user_folder, filename)
        face_file.save(image_path)
        
        # Process with face_recognition
        image = face_recognition.load_image_file(image_path)
        face_encodings = face_recognition.face_encodings(image)
        
        if not face_encodings:
            os.remove(image_path)
            conn.close()
            return jsonify({'success': False, 'message': 'Tidak ada wajah terdeteksi dalam gambar'})
        
        if len(face_encodings) > 1:
            os.remove(image_path)
            conn.close()
            return jsonify({'success': False, 'message': 'Terdeteksi lebih dari satu wajah. Gunakan foto dengan satu wajah saja'})
        
        # Get the face encoding
        face_encoding = face_encodings[0]
        encoding_json = json.dumps(face_encoding.tolist())
        
        # Deactivate old face data
        conn.execute('UPDATE face_data SET active = 0 WHERE user_id = ?', (session['user_id'],))
        
        # Save new encoding to database
        conn.execute('''
            INSERT INTO face_data (user_id, face_encoding, photo_path, active)
            VALUES (?, ?, ?, 1)
        ''', (session['user_id'], encoding_json, image_path))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Face recognition berhasil disetup!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/remove_face', methods=['POST'])
@login_required
def remove_face():
    """Remove face recognition for user"""
    try:
        conn = get_db_connection()
        
        # Get face data to remove files
        face_data = conn.execute(
            'SELECT photo_path FROM face_data WHERE user_id = ? AND active = 1',
            (session['user_id'],)
        ).fetchall()
        
        # Remove face data from database
        conn.execute('UPDATE face_data SET active = 0 WHERE user_id = ?', (session['user_id'],))
        conn.commit()
        conn.close()
        
        # Remove physical files
        for data in face_data:
            if data['photo_path'] and os.path.exists(data['photo_path']):
                try:
                    os.remove(data['photo_path'])
                except Exception:
                    pass  # Continue even if file removal fails
        
        return jsonify({'success': True, 'message': 'Face recognition berhasil dihapus!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/delete_coordinate', methods=['POST'])
@login_required
def delete_coordinate():
    """Delete coordinate with enhanced error handling"""
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
    
    try:
        coordinate_id = request.form.get('id')
        print(f"🔍 DEBUG: Attempting to delete coordinate ID: {coordinate_id}")
        
        if not coordinate_id:
            return jsonify({'success': False, 'message': 'ID koordinat tidak ditemukan'}), 400
        
        try:
            coordinate_id = int(coordinate_id)
        except ValueError:
            return jsonify({'success': False, 'message': 'ID koordinat tidak valid'}), 400
            
        conn = get_db_connection()
        
        # Check if coordinate exists
        coordinate = conn.execute('SELECT * FROM coordinates WHERE id = ?', (coordinate_id,)).fetchone()
        if not coordinate:
            conn.close()
            return jsonify({'success': False, 'message': 'Koordinat tidak ditemukan'}), 404
        
        coordinate_name = coordinate['name']
        print(f"✅ DEBUG: Found coordinate: {coordinate_name}")
        
        # Delete the coordinate with explicit transaction
        conn.execute('BEGIN IMMEDIATE')
        try:
            # Perform deletion
            cursor = conn.execute('DELETE FROM coordinates WHERE id = ?', (coordinate_id,))
            deleted_rows = cursor.rowcount
            
            if deleted_rows > 0:
                
                conn.commit()
                print(f"💾 DEBUG: Successfully deleted {deleted_rows} row(s)")
                
                check = conn.execute('SELECT COUNT(*) as count FROM coordinates WHERE id = ?', (coordinate_id,)).fetchone()
                
                if check['count'] == 0:
                    conn.close()
                    return jsonify({
                        'success': True,
                        'message': f'Koordinat "{coordinate_name}" berhasil dihapus!',
                        'deleted_id': coordinate_id,
                        'reload_required': True
                    })
                else:
                    conn.rollback()
                    conn.close()
                    return jsonify({'success': False, 'message': 'Gagal menghapus - data masih ada'}), 500
            else:
                conn.rollback()
                conn.close()
                return jsonify({'success': False, 'message': 'Tidak ada data yang dihapus'}), 500
                
        except Exception as transaction_error:
            conn.rollback()
            conn.close()
            print(f"💥 DEBUG: Transaction error: {str(transaction_error)}")
            raise transaction_error
            
    except Exception as e:
        print(f"💥 DEBUG: Exception occurred: {str(e)}")
        print(f"📋 DEBUG: Exception type: {type(e).__name__}")
        import traceback
        print(f"🔥 DEBUG: Traceback: {traceback.format_exc()}")
        
        if 'conn' in locals():
            try:
                conn.rollback()
                conn.close()
            except:
                pass
                
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/api/coordinates/list', methods=['GET'])
@login_required
def api_coordinates_list():
    """API to get fresh coordinates list"""
    if session.get('username') != 'admin':
        return jsonify({
            'success': False, 
            'message': 'Access denied'
        }), 403
    
    try:
        conn = get_db_connection()
        
        coordinates_raw = conn.execute(
            'SELECT * FROM coordinates ORDER BY id DESC'
        ).fetchall()
        
        # Convert to list of dicts
        coordinates_list = []
        for coord in coordinates_raw:
            coordinates_list.append({
                'id': coord['id'],
                'name': coord['name'],
                'latitude': coord['latitude'],
                'longitude': coord['longitude'],
                'radius': coord['radius'],
                'active': coord['active']
            })
        
        conn.close()
        
        print(f"✅ API: Returning {len(coordinates_list)} coordinates")
        
        return jsonify({
            'success': True,
            'coordinates': coordinates_list,
            'count': len(coordinates_list)
        })
        
    except Exception as e:
        print(f"❌ API Error: {e}")
        return jsonify({
            'success': False,
            'message': str(e),
            'coordinates': []
        }), 500
        
@app.route('/set_coordinat')
@login_required
def set_coordinat():
    """Coordinate settings page (admin only)"""
    if session.get('username') != 'admin':
        flash('Access denied. Admin only.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    try:

        coordinates_raw = conn.execute(
            'SELECT * FROM coordinates ORDER BY id DESC'
        ).fetchall()
        
        # Convert Row objects to dict
        coordinates = []
        for row in coordinates_raw:
            coord_dict = {
                'id': row['id'],
                'name': row['name'],
                'latitude': row['latitude'],
                'longitude': row['longitude'],
                'radius': row['radius'],
                'active': row['active']
            }
            coordinates.append(coord_dict)
            
        print(f"✅ Loaded {len(coordinates)} coordinates for page render")
        
        # Debug: Print untuk troubleshooting
        if coordinates:
            print(f"📍 First coordinate: {coordinates[0]}")
        else:
            print("⚠️ No coordinates found in database")
        
        # ✅ TAMBAHKAN: Get user info untuk navbar
        user = conn.execute(
            'SELECT id, username, full_name, role FROM users WHERE id = ?', 
            (session['user_id'],)
        ).fetchone()
        
        # ✅ TAMBAHKAN: Store user info in session if not exists
        if user and not session.get('full_name'):
            session['full_name'] = user['full_name']
            session['role'] = user['role']
        
    except Exception as e:
        print(f"❌ Error loading coordinates: {e}")
        coordinates = []
        user = None
    finally:
        conn.close()
    
    return render_template('set_coordinat.html', 
                         coordinates=coordinates,
                         user=user)

@app.route('/add_coordinate', methods=['POST'])
@login_required
def add_coordinate():
    """Add new coordinate - Returns JSON for AJAX"""
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
    
    try:
        name = request.form.get('name', '').strip()
        latitude = request.form.get('latitude')
        longitude = request.form.get('longitude')
        radius = request.form.get('radius', 100)
        
        # Validasi input
        if not name or not latitude or not longitude:
            return jsonify({'success': False, 'message': 'Data tidak lengkap'}), 400
        
        # Convert to proper types
        latitude = float(latitude)
        longitude = float(longitude)
        radius = int(radius)
        
        # Validasi range
        if latitude < -90 or latitude > 90:
            return jsonify({'success': False, 'message': 'Latitude harus antara -90 dan 90'}), 400
        
        if longitude < -180 or longitude > 180:
            return jsonify({'success': False, 'message': 'Longitude harus antara -180 dan 180'}), 400
        
        if radius < 10 or radius > 1000:
            return jsonify({'success': False, 'message': 'Radius harus antara 10 dan 1000 meter'}), 400
        
        conn = get_db_connection()
        
        existing = conn.execute(
            'SELECT id, name FROM coordinates WHERE LOWER(name) = LOWER(?)', 
            (name,)
        ).fetchone()
        
        if existing:
            conn.close()
            return jsonify({
                'success': False, 
                'message': f'Lokasi dengan nama "{existing["name"]}" sudah ada!'
            }), 400
        
        # Insert new coordinate
        cursor = conn.execute(
            'INSERT INTO coordinates (name, latitude, longitude, radius, active) VALUES (?, ?, ?, ?, 1)',
            (name, latitude, longitude, radius)
        )
        
        new_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        print(f"✅ Added new coordinate: {name} (ID: {new_id})")
        
        return jsonify({
            'success': True,
            'message': 'Koordinat berhasil ditambahkan!',
            'coordinate': {
                'id': new_id,
                'name': name,
                'latitude': latitude,
                'longitude': longitude,
                'radius': radius
            }
        })
        
    except ValueError as e:
        return jsonify({'success': False, 'message': 'Format data tidak valid'}), 400
    except Exception as e:
        print(f"❌ Error adding coordinate: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500    
@app.route('/toggle_coordinate_status', methods=['POST'])
@login_required
def toggle_coordinate_status():
    """Toggle coordinate active status"""
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
    
    try:
        coordinate_id = request.form.get('id')
        
        if not coordinate_id:
            return jsonify({'success': False, 'message': 'ID koordinat tidak ditemukan'}), 400
        
        conn = get_db_connection()
        
        # Get current status
        coordinate = conn.execute('SELECT * FROM coordinates WHERE id = ?', (coordinate_id,)).fetchone()
        if not coordinate:
            conn.close()
            return jsonify({'success': False, 'message': 'Koordinat tidak ditemukan'}), 404
        
        # Toggle status
        new_status = not coordinate['active']
        conn.execute('UPDATE coordinates SET active = ? WHERE id = ?', (new_status, coordinate_id))
        conn.commit()
        conn.close()
        
        status_text = 'diaktifkan' if new_status else 'dinonaktifkan'
        
        return jsonify({
            'success': True,
            'message': f'Koordinat "{coordinate["name"]}" berhasil {status_text}!',
            'new_status': new_status
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/update_coordinate', methods=['POST'])
@login_required
def update_coordinate():
    """Update coordinate"""
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
    
    try:
        coordinate_id = request.form.get('id')
        name = request.form.get('name')
        latitude = float(request.form.get('latitude'))
        longitude = float(request.form.get('longitude'))
        radius = int(request.form.get('radius', 100))
        
        if not all([coordinate_id, name]):
            return jsonify({'success': False, 'message': 'Data tidak lengkap'}), 400
        
        conn = get_db_connection()
        
        # Check if coordinate exists
        coordinate = conn.execute('SELECT * FROM coordinates WHERE id = ?', (coordinate_id,)).fetchone()
        if not coordinate:
            conn.close()
            return jsonify({'success': False, 'message': 'Koordinat tidak ditemukan'}), 404
        
        # Update coordinate
        conn.execute(
            '''UPDATE coordinates 
               SET name = ?, latitude = ?, longitude = ?, radius = ?
               WHERE id = ?''',
            (name, latitude, longitude, radius, coordinate_id)
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Koordinat "{name}" berhasil diperbarui!'
        })
        
    except ValueError as e:
        return jsonify({'success': False, 'message': 'Format data tidak valid'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/logout')
def logout():
    """Logout"""
    session.clear()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('login'))

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500


@app.route('/api/attendance/monthly', methods=['GET'])
@login_required
def api_monthly_attendance():
    """API to get monthly attendance data"""
    try:
        # Get month and year from query params, default to current month
        from datetime import datetime, timedelta
        import calendar
        
        month = request.args.get('month', datetime.now().month, type=int)
        year = request.args.get('year', datetime.now().year, type=int)
        
        # Validate month and year
        if month < 1 or month > 12:
            month = datetime.now().month
        if year < 2020 or year > 2030:
            year = datetime.now().year
            
        # Get first and last day of the month
        first_day = datetime(year, month, 1).strftime('%Y-%m-%d')
        last_day = datetime(year, month, calendar.monthrange(year, month)[1]).strftime('%Y-%m-%d')
        
        conn = get_db_connection()
        
        # Get attendance data for the month
        attendance_data = conn.execute('''
            SELECT 
                date,
                time_in,
                time_out,
                latitude,
                longitude,
                photo_path,
                CASE 
                    WHEN time_in IS NOT NULL AND time_out IS NOT NULL THEN 'complete'
                    WHEN time_in IS NOT NULL AND time_out IS NULL THEN 'incomplete'
                    ELSE 'absent'
                END as status,
                CASE 
                    WHEN time_in IS NOT NULL AND time_out IS NOT NULL THEN
                        CAST((julianday(date || ' ' || time_out) - julianday(date || ' ' || time_in)) * 24 * 60 AS INTEGER)
                    ELSE NULL
                END as work_minutes
            FROM attendance 
            WHERE user_id = ? AND date BETWEEN ? AND ?
            ORDER BY date DESC
        ''', (session['user_id'], first_day, last_day)).fetchall()
        
        # Convert to list of dictionaries
        attendance_list = []
        for row in attendance_data:
            attendance_list.append({
                'date': row['date'],
                'time_in': row['time_in'],
                'time_out': row['time_out'],
                'status': row['status'],
                'work_minutes': row['work_minutes'],
                'work_hours': round(row['work_minutes'] / 60, 1) if row['work_minutes'] else None,
                'has_photo': bool(row['photo_path'])
            })
        
        # Calculate statistics
        total_days = len(attendance_list)
        present_days = len([a for a in attendance_list if a['status'] in ['complete', 'incomplete']])
        complete_days = len([a for a in attendance_list if a['status'] == 'complete'])
        incomplete_days = len([a for a in attendance_list if a['status'] == 'incomplete'])
        absent_days = total_days - present_days
        
        # Calculate total work hours
        total_work_minutes = sum([a['work_minutes'] for a in attendance_list if a['work_minutes']])
        total_work_hours = round(total_work_minutes / 60, 1) if total_work_minutes else 0
        avg_work_hours = round(total_work_hours / complete_days, 1) if complete_days > 0 else 0
        
        # Attendance rate
        working_days = calendar.monthrange(year, month)[1]  # Total days in month
        # You might want to exclude weekends, but for simplicity we'll use all days
        attendance_rate = round((present_days / working_days) * 100, 1) if working_days > 0 else 0
        
        stats = {
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month],
            'total_days': working_days,
            'present_days': present_days,
            'complete_days': complete_days,
            'incomplete_days': incomplete_days,
            'absent_days': working_days - present_days,
            'attendance_rate': attendance_rate,
            'total_work_hours': total_work_hours,
            'avg_work_hours': avg_work_hours
        }
        
        conn.close()
        
        return jsonify({
            'success': True,
            'attendance': attendance_list,
            'stats': stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/users')
@login_required
def users_dashboard():
    """Simple users dashboard"""
    conn = get_db_connection()
    classes = conn.execute('SELECT * FROM classes WHERE active = 1 ORDER BY name').fetchall()
    conn.close()
    return render_template('users_dashboard.html', classes=classes)

@app.route('/api/users/list', methods=['GET'])
@login_required
def api_users_list():
    """API to get list of users - accessible by all logged in users"""
    try:
        conn = get_db_connection()
        
        # Get basic user information (tidak termasuk info sensitif)
        users = conn.execute('''
            SELECT 
                u.id, u.username, u.full_name, u.role, u.active,
                COALESCE(c.name, '-') as class_name, 
                u.created_at, u.updated_at,
                CASE WHEN f.id IS NOT NULL THEN 1 ELSE 0 END as face_recognition,
                a.last_attendance
            FROM users u
            LEFT JOIN classes c ON u.class_id = c.id
            LEFT JOIN (
                SELECT user_id, MAX(id) as id 
                FROM face_data 
                WHERE active = 1 
                GROUP BY user_id
            ) f ON u.id = f.user_id
            LEFT JOIN (
                SELECT user_id, MAX(date) as last_attendance
                FROM attendance
                WHERE time_in IS NOT NULL
                GROUP BY user_id
            ) a ON u.id = a.user_id
            ORDER BY u.full_name ASC
        ''').fetchall()
        
        users_list = []
        for user in users:
            users_list.append({
                'id': user['id'],
                'username': user['username'],
                'full_name': user['full_name'],
                'class_name': user['class_name'],
                'role': user['role'],
                'active': user['active'],
                'face_recognition': user['face_recognition'],
                'created_at': user['created_at'],
                'updated_at': user['updated_at'],
                'last_attendance': user['last_attendance']
            })
        
        # Get statistics
        today = datetime.now().strftime("%Y-%m-%d")
        
        # Total users
        total_users = len(users_list)
        
        # Active users
        active_users = len([u for u in users_list if u['active']])
        
        # Face recognition enabled users
        face_enabled_users = len([u for u in users_list if u['face_recognition']])
        
        # Today attendance count
        today_attendance = conn.execute('''
            SELECT COUNT(DISTINCT user_id) as count
            FROM attendance 
            WHERE date = ? AND time_in IS NOT NULL
        ''', (today,)).fetchone()['count']
        
        stats = {
            'total_users': total_users,
            'active_users': active_users,
            'face_enabled_users': face_enabled_users,
            'today_attendance': today_attendance
        }
        
        conn.close()
        return jsonify({
            'success': True,
            'users': users_list,
            'stats': stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        
@app.route('/api/users/create', methods=['POST'])
@login_required
def api_create_user():
    """API to create new user"""
    try:
        
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['username', 'full_name', 'class_id', 'password']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False, 
                    'message': f'Field {field} is required'
                }), 400
        
        # Validate password length
        if len(data.get('password', '')) < 6:
            return jsonify({
                'success': False,
                'message': 'Password must be at least 6 characters long'
            }), 400
        
        conn = get_db_connection()
        
        # Check if username already exists
        existing = conn.execute(
            'SELECT id FROM users WHERE username = ?',
            (data['username'],)
        ).fetchone()
        
        if existing:
            conn.close()
            return jsonify({
                'success': False,
                'message': 'Username already exists'
            }), 400
        
        # Hash password
        hashed_password = generate_password_hash(data['password'])
        
        # Insert new user
        conn.execute('''
            INSERT INTO users (username, full_name, class_id, password, role, active)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            data['username'],
            data['full_name'],
            int(data['class_id']),
            hashed_password,
            data.get('role', 'user'),
            data.get('active', True)
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'User created successfully'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error creating user: {str(e)}'
        }), 500

@app.route('/api/users/update/<int:user_id>', methods=['PUT'])
@login_required
def api_update_user(user_id):
    """API to update user"""
    try:

        data = request.get_json()
        
        # Validate required fields
        required_fields = ['full_name']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'Field {field} is required'
                }), 400
        
        conn = get_db_connection()
        
        # Check if user exists
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if not user:
            conn.close()
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 404
        
        # Prepare update query
        update_fields = []
        update_values = []
        
        # Update full_name
        update_fields.append('full_name = ?')
        update_values.append(data['full_name'])
        
        # Update class_id if provided
        if 'class_id' in data:
            update_fields.append('class_id = ?')
            update_values.append(int(data['class_id']))
        
        # Update role if provided
        if 'role' in data:
            update_fields.append('role = ?')
            update_values.append(data['role'])
        
        # Update active status if provided
        if 'active' in data:
            update_fields.append('active = ?')
            update_values.append(data['active'])
        
        # Update password if provided
        if data.get('password'):
            if len(data['password']) < 6:
                conn.close()
                return jsonify({
                    'success': False,
                    'message': 'Password must be at least 6 characters long'
                }), 400
            
            update_fields.append('password = ?')
            update_values.append(generate_password_hash(data['password']))
        
        # Add updated_at
        update_fields.append('updated_at = CURRENT_TIMESTAMP')
        
        # Add user_id for WHERE clause
        update_values.append(user_id)
        
        # Execute update
        query = f"UPDATE users SET {', '.join(update_fields)} WHERE id = ?"
        conn.execute(query, update_values)
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'User updated successfully'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error updating user: {str(e)}'
        }), 500

@app.route('/api/users/delete/<int:user_id>', methods=['DELETE'])
@login_required
def api_delete_user(user_id):
    """API to delete user"""
    try:
        # Check if user has admin privileges
        # if session.get('username') != 'admin':
        #     return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
        
        # Prevent self-deletion
        if session.get('user_id') == user_id:
            return jsonify({
                'success': False,
                'message': 'You cannot delete your own account'
            }), 400
        
        conn = get_db_connection()
        
        # Check if user exists
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if not user:
            conn.close()
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 404
        
        # Get face data to clean up files
        face_data = conn.execute(
            'SELECT photo_path FROM face_data WHERE user_id = ?',
            (user_id,)
        ).fetchall()
        
        # Delete related data first (to maintain referential integrity)
        # Delete attendance records
        conn.execute('DELETE FROM attendance WHERE user_id = ?', (user_id,))
        
        # Delete attendance logs
        conn.execute('DELETE FROM attendance_logs WHERE user_id = ?', (user_id,))
        
        # Delete face data
        conn.execute('DELETE FROM face_data WHERE user_id = ?', (user_id,))
        
        # Delete user
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        
        conn.commit()
        conn.close()
        
        # Clean up face recognition files
        for data in face_data:
            if data['photo_path'] and os.path.exists(data['photo_path']):
                try:
                    os.remove(data['photo_path'])
                    # Also try to remove the directory if it's empty
                    dir_path = os.path.dirname(data['photo_path'])
                    if os.path.exists(dir_path) and not os.listdir(dir_path):
                        os.rmdir(dir_path)
                except Exception:
                    pass 
        
        return jsonify({
            'success': True,
            'message': 'User deleted successfully'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error deleting user: {str(e)}'
        }), 500

@app.route('/api/users/detail/<int:user_id>', methods=['GET'])
@login_required
def api_user_detail(user_id):
    """API to get user detail"""
    try:
        conn = get_db_connection()
        
        # Get user with additional info
        user = conn.execute('''
            SELECT 
                u.*,
                CASE WHEN f.id IS NOT NULL THEN 1 ELSE 0 END as face_recognition,
                COUNT(DISTINCT a.id) as total_attendance,
                MAX(a.date) as last_attendance_date,
                COUNT(DISTINCT CASE WHEN a.date >= date('now', 'start of month') THEN a.id END) as this_month_attendance
            FROM users u
            LEFT JOIN face_data f ON u.id = f.user_id AND f.active = 1
            LEFT JOIN attendance a ON u.id = a.user_id AND a.time_in IS NOT NULL
            WHERE u.id = ?
            GROUP BY u.id
        ''', (user_id,)).fetchone()
        
        if not user:
            conn.close()
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 404
        
        # Get recent attendance (last 10 records)
        recent_attendance = conn.execute('''
            SELECT date, time_in, time_out
            FROM attendance
            WHERE user_id = ? AND time_in IS NOT NULL
            ORDER BY date DESC
            LIMIT 10
        ''', (user_id,)).fetchall()
        
        conn.close()
        
        user_dict = dict(user)
        user_dict['recent_attendance'] = [dict(row) for row in recent_attendance]
        
        return jsonify({
            'success': True,
            'user': user_dict
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting user detail: {str(e)}'
        }), 500

@app.route('/api/users/toggle-status/<int:user_id>', methods=['POST'])
@login_required
def api_toggle_user_status(user_id):
    """API to toggle user active status"""
    try:
        # Check if user has admin privileges
        # if session.get('username') != 'admin':
        #     return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
        
        # Prevent self-deactivation
        if session.get('user_id') == user_id:
            return jsonify({
                'success': False,
                'message': 'You cannot deactivate your own account'
            }), 400
        
        conn = get_db_connection()
        
        # Get current status
        user = conn.execute('SELECT active FROM users WHERE id = ?', (user_id,)).fetchone()
        if not user:
            conn.close()
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 404
        
        # Toggle status
        new_status = not user['active']
        conn.execute(
            'UPDATE users SET active = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
            (new_status, user_id)
        )
        
        conn.commit()
        conn.close()
        
        status_text = 'activated' if new_status else 'deactivated'
        
        return jsonify({
            'success': True,
            'message': f'User {status_text} successfully',
            'new_status': new_status
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error toggling user status: {str(e)}'
        }), 500

# Optional: Bulk operations
@app.route('/api/users/bulk-delete', methods=['POST'])
@login_required
def api_bulk_delete_users():
    """API to delete multiple users"""
    try:
        # Check if user has admin privileges
        # if session.get('username') != 'admin':
        #     return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
        
        data = request.get_json()
        user_ids = data.get('user_ids', [])
        
        if not user_ids:
            return jsonify({
                'success': False,
                'message': 'No users selected'
            }), 400
        
        # Prevent self-deletion
        if session.get('user_id') in user_ids:
            return jsonify({
                'success': False,
                'message': 'You cannot delete your own account'
            }), 400
        
        conn = get_db_connection()
        
        deleted_count = 0
        for user_id in user_ids:
            # Get face data for cleanup
            face_data = conn.execute(
                'SELECT photo_path FROM face_data WHERE user_id = ?',
                (user_id,)
            ).fetchall()
            
            # Delete related data
            conn.execute('DELETE FROM attendance WHERE user_id = ?', (user_id,))
            conn.execute('DELETE FROM attendance_logs WHERE user_id = ?', (user_id,))
            conn.execute('DELETE FROM face_data WHERE user_id = ?', (user_id,))
            
            # Delete user
            result = conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
            if result.rowcount > 0:
                deleted_count += 1
                
                # Cleanup files
                for data in face_data:
                    if data['photo_path'] and os.path.exists(data['photo_path']):
                        try:
                            os.remove(data['photo_path'])
                        except Exception:
                            pass
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{deleted_count} users deleted successfully'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error deleting users: {str(e)}'
        }), 500
        
# Tambahkan endpoint ini ke file app.py

@app.route('/api/attendance/daily', methods=['GET'])
@login_required
def api_daily_attendance():
    """API to get daily attendance data"""
    try:
        # Get date from query params, default to today
        from datetime import datetime
        
        date_param = request.args.get('date')
        if date_param:
            try:
                # Validate date format
                selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Invalid date format. Use YYYY-MM-DD'
                }), 400
        else:
            selected_date = datetime.now().date()
        
        date_str = selected_date.strftime('%Y-%m-%d')
        
        conn = get_db_connection()
        
        # Get attendance data for the selected date with user information
        attendance_data = conn.execute('''
            SELECT 
                a.id,
                a.user_id,
                a.date,
                a.time_in,
                a.time_out,
                a.latitude,
                a.longitude,
                a.photo_path,
                u.username,
                u.full_name,
                COALESCE(c.name, '-') as class_name,
                CASE 
                    WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN 'complete'
                    WHEN a.time_in IS NOT NULL AND a.time_out IS NULL THEN 'incomplete'
                    ELSE 'absent'
                END as status,
                CASE 
                    WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN
                        CAST((julianday(a.date || ' ' || a.time_out) - julianday(a.date || ' ' || a.time_in)) * 24 * 60 AS INTEGER)
                    ELSE NULL
                END as work_minutes
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            LEFT JOIN classes c ON u.class_id = c.id
            WHERE a.date = ? AND a.time_in IS NOT NULL AND u.role != 'admin'
            ORDER BY a.time_in ASC
        ''', (date_str,)).fetchall()
        
        # Convert to list of dictionaries
        attendance_list = []
        for row in attendance_data:
            attendance_list.append({
                'id': row['id'],
                'user_id': row['user_id'],
                'username': row['username'],
                'full_name': row['full_name'],
                'class_name': row['class_name'],    
                'date': row['date'],
                'time_in': row['time_in'],
                'time_out': row['time_out'],
                'status': row['status'],
                'work_minutes': row['work_minutes'],
                'work_hours': round(row['work_minutes'] / 60, 1) if row['work_minutes'] else None,
                'has_photo': bool(row['photo_path'])
            })
        
        # Calculate daily statistics
        total_present = len(attendance_list)
        complete_attendance = len([a for a in attendance_list if a['status'] == 'complete'])
        incomplete_attendance = len([a for a in attendance_list if a['status'] == 'incomplete'])
        
        # Calculate total work hours for the day
        total_work_minutes = sum([a['work_minutes'] for a in attendance_list if a['work_minutes']])
        total_work_hours = round(total_work_minutes / 60, 1) if total_work_minutes else 0
        avg_work_hours = round(total_work_hours / complete_attendance, 1) if complete_attendance > 0 else 0
        
        # Get total registered users for attendance rate calculation
        total_users = conn.execute('SELECT COUNT(*) as count FROM users WHERE active = 1').fetchone()['count']
        attendance_rate = round((total_present / total_users) * 100, 1) if total_users > 0 else 0
        
        daily_stats = {
            'date': date_str,
            'total_present': total_present,
            'complete_attendance': complete_attendance,
            'incomplete_attendance': incomplete_attendance,
            'total_registered_users': total_users,
            'attendance_rate': attendance_rate,
            'total_work_hours': total_work_hours,
            'avg_work_hours': avg_work_hours
        }
        
        conn.close()
        
        return jsonify({
            'success': True,
            'date': date_str,
            'attendance': attendance_list,
            'stats': daily_stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/attendance/weekly', methods=['GET'])
@login_required
def api_weekly_attendance():
    """API to get weekly attendance summary"""
    try:
        from datetime import datetime, timedelta
        
        # Get week start date (default to current week)
        week_start_param = request.args.get('week_start')
        if week_start_param:
            try:
                week_start = datetime.strptime(week_start_param, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Invalid date format. Use YYYY-MM-DD'
                }), 400
        else:
            today = datetime.now().date()
            # Get Monday of current week
            week_start = today - timedelta(days=today.weekday())
        
        # Calculate week end (Sunday)
        week_end = week_start + timedelta(days=6)
        
        conn = get_db_connection()
        
        # Get attendance summary for each day of the week
        weekly_data = []
        current_date = week_start
        
        while current_date <= week_end:
            date_str = current_date.strftime('%Y-%m-%d')
            
            # Get attendance count for this date
            daily_count = conn.execute('''
                SELECT 
                    COUNT(*) as total_present,
                    SUM(CASE WHEN time_out IS NOT NULL THEN 1 ELSE 0 END) as complete_count
                FROM attendance 
                WHERE date = ? AND time_in IS NOT NULL
            ''', (date_str,)).fetchone()
            
            # Get total active users
            total_users = conn.execute('SELECT COUNT(*) FROM users WHERE active = 1').fetchone()[0]
            
            attendance_rate = round((daily_count['total_present'] / total_users) * 100, 1) if total_users > 0 else 0
            
            weekly_data.append({
                'date': date_str,
                'day_name': current_date.strftime('%A'),
                'total_present': daily_count['total_present'],
                'complete_count': daily_count['complete_count'],
                'incomplete_count': daily_count['total_present'] - daily_count['complete_count'],
                'total_users': total_users,
                'attendance_rate': attendance_rate
            })
            
            current_date += timedelta(days=1)
        
        # Calculate weekly summary
        total_weekly_present = sum([day['total_present'] for day in weekly_data])
        total_weekly_complete = sum([day['complete_count'] for day in weekly_data])
        avg_daily_attendance = round(total_weekly_present / 7, 1)
        avg_attendance_rate = round(sum([day['attendance_rate'] for day in weekly_data]) / 7, 1)
        
        weekly_summary = {
            'week_start': week_start.strftime('%Y-%m-%d'),
            'week_end': week_end.strftime('%Y-%m-%d'),
            'total_weekly_present': total_weekly_present,
            'total_weekly_complete': total_weekly_complete,
            'avg_daily_attendance': avg_daily_attendance,
            'avg_attendance_rate': avg_attendance_rate
        }
        
        conn.close()
        
        return jsonify({
            'success': True,
            'weekly_data': weekly_data,
            'summary': weekly_summary
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration page"""
    conn = get_db_connection()
    
    #query classes untuk dropdown
    classes = conn.execute('SELECT * FROM classes WHERE active = 1 ORDER BY name').fetchall()
    
    if request.method == 'POST':
        username = request.form.get('username')
        full_name = request.form.get('full_name')
        class_id = request.form.get('class_id')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        # Validation
        if not all([username, full_name, class_id, password, confirm_password]):
            flash('Semua field wajib harus diisi!', 'error')
            conn.close()
            return render_template('register.html', classes=classes)
        
        if len(username) < 3:
            flash('Username minimal 3 karakter!', 'error')
            conn.close()
            return render_template('register.html', classes=classes)
        
        if len(password) < 6:
            flash('Password minimal 6 karakter!', 'error')
            conn.close()
            return render_template('register.html', classes=classes)
        
        if password != confirm_password:
            flash('Password dan konfirmasi password tidak cocok!', 'error')
            conn.close()
            return render_template('register.html', classes=classes)
        
        if not any(c.isalpha() for c in password) or not any(c.isdigit() for c in password):
            flash('Password harus mengandung huruf dan angka!', 'error')
            conn.close()
            return render_template('register.html', classes=classes)
        
        # Check if username exists
        existing_user = conn.execute(
            'SELECT id FROM users WHERE username = ?', (username,)
        ).fetchone()
        
        if existing_user:
            conn.close()
            flash('Username sudah digunakan!', 'error')
            return render_template('register.html', classes=classes)
        
        try:
            hashed_password = generate_password_hash(password)
            
            # ✅ Insert dengan class_id
            conn.execute('''
                INSERT INTO users (username, full_name, class_id, password, role, active)
                VALUES (?, ?, ?, ?, 'user', 1)
            ''', (username, full_name, int(class_id), hashed_password))
            
            conn.commit()
            conn.close()
            
            flash('Registrasi Berhasil!', 'success')
            session['show_face_setup_reminder'] = True
            
            return redirect(url_for('register'))
            
        except Exception as e:
            conn.close()
            flash(f'Error saat registrasi: {str(e)}', 'error')
            return render_template('register.html', classes=classes)
    
    conn.close()
    return render_template('register.html', classes=classes)

# API untuk check username availability
@app.route('/api/check_username', methods=['POST'])
def api_check_username():
    """API to check username availability"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        
        if not username:
            return jsonify({'available': False, 'message': 'Username tidak boleh kosong'})
        
        if len(username) < 3:
            return jsonify({'available': False, 'message': 'Username minimal 3 karakter'})
        
        conn = get_db_connection()
        existing = conn.execute(
            'SELECT id FROM users WHERE username = ?', (username,)
        ).fetchone()
        conn.close()
        
        if existing:
            return jsonify({'available': False, 'message': 'Username sudah digunakan'})
        else:
            return jsonify({'available': True, 'message': 'Username tersedia'})
    
    except Exception as e:
        return jsonify({'available': False, 'message': 'Error checking username'})

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        user = conn.execute(
            'SELECT * FROM users WHERE username = ?', (username,)
        ).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['full_name'] = user['full_name']
            
            # Check if user should see face setup reminder
            if session.get('show_face_setup_reminder'):
                session.pop('show_face_setup_reminder', None)
                session['show_face_reminder_on_dashboard'] = True
            
            # flash('Login berhasil!', 'success')
            
            # Redirect based on role
            if username == 'admin':
                return redirect(url_for('users_dashboard'))
            else:
                return redirect(url_for('index'))
        else:
            flash('Username atau password salah!', 'error')
    
    return render_template('login.html')

@app.route('/api/set_face_reminder', methods=['POST'])
@login_required
def api_set_face_reminder():
    """API to set face recognition reminder for next login"""
    try:

        return jsonify({
            'success': True,
            'message': 'Reminder set for next login'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })
        
        
@app.route('/api/export/users', methods=['GET'])
@login_required
def export_users_excel():
    """Export users data to Excel grouped by class"""
    try:
        conn = get_db_connection()
        
        # Get all classes
        classes = conn.execute('SELECT * FROM classes WHERE active = 1 ORDER BY name').fetchall()
        
        # Get users without class
        users_no_class = conn.execute('''
            SELECT 
                u.id as "ID",
                u.username as "Username",
                u.full_name as "Nama Lengkap",
                u.role as "Role",
                CASE WHEN u.active = 1 THEN 'Aktif' ELSE 'Nonaktif' END as "Status",
                CASE WHEN f.id IS NOT NULL THEN 'Ya' ELSE 'Tidak' END as "Face Recognition",
                u.created_at as "Tanggal Daftar",
                a.last_attendance as "Terakhir Hadir",
                COUNT(att.id) as "Total Kehadiran"
            FROM users u
            LEFT JOIN (
                SELECT user_id, MAX(id) as id 
                FROM face_data 
                WHERE active = 1 
                GROUP BY user_id
            ) f ON u.id = f.user_id
            LEFT JOIN (
                SELECT user_id, MAX(date) as last_attendance
                FROM attendance
                WHERE time_in IS NOT NULL
                GROUP BY user_id
            ) a ON u.id = a.user_id
            LEFT JOIN attendance att ON u.id = att.user_id AND att.time_in IS NOT NULL
            WHERE u.role != 'admin' AND (u.class_id IS NULL OR u.class_id = 0)
            GROUP BY u.id
            ORDER BY u.full_name ASC
        ''').fetchall()
        
        conn.close()
        
        # Create Excel file with multiple sheets
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Create summary sheet first
            summary_data = []
            total_all_students = 0
            
            # Process each class
            for cls in classes:
                conn = get_db_connection()
                
                users_in_class = conn.execute('''
                    SELECT 
                        u.id as "ID",
                        u.username as "Username",
                        u.full_name as "Nama Lengkap",
                        u.role as "Role",
                        CASE WHEN u.active = 1 THEN 'Aktif' ELSE 'Nonaktif' END as "Status",
                        CASE WHEN f.id IS NOT NULL THEN 'Ya' ELSE 'Tidak' END as "Face Recognition",
                        u.created_at as "Tanggal Daftar",
                        a.last_attendance as "Terakhir Hadir",
                        COUNT(att.id) as "Total Kehadiran"
                    FROM users u
                    LEFT JOIN (
                        SELECT user_id, MAX(id) as id 
                        FROM face_data 
                        WHERE active = 1 
                        GROUP BY user_id
                    ) f ON u.id = f.user_id
                    LEFT JOIN (
                        SELECT user_id, MAX(date) as last_attendance
                        FROM attendance
                        WHERE time_in IS NOT NULL
                        GROUP BY user_id
                    ) a ON u.id = a.user_id
                    LEFT JOIN attendance att ON u.id = att.user_id AND att.time_in IS NOT NULL
                    WHERE u.role != 'admin' AND u.class_id = ?
                    GROUP BY u.id
                    ORDER BY u.full_name ASC
                ''', (cls['id'],)).fetchall()
                
                conn.close()
                
                if len(users_in_class) > 0:
                    # Convert to DataFrame
                    class_data = [dict(user) for user in users_in_class]
                    df_class = pd.DataFrame(class_data)
                    
                    # Format dates
                    date_columns = ['Tanggal Daftar', 'Terakhir Hadir']
                    for col in date_columns:
                        if col in df_class.columns:
                            df_class[col] = pd.to_datetime(df_class[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M')
                    
                    # Sanitize sheet name (max 31 chars, no special chars)
                    sheet_name = cls['name'][:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '')
                    
                    # Write to sheet
                    df_class.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust columns
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add to summary
                    total_all_students += len(users_in_class)
                    active_count = len([u for u in class_data if u['Status'] == 'Aktif'])
                    face_count = len([u for u in class_data if u['Face Recognition'] == 'Ya'])
                    
                    summary_data.append({
                        'Kelas': cls['name'],
                        'Jumlah Siswa': len(users_in_class),
                        'Aktif': active_count,
                        'Face Recognition': face_count
                    })
            
            # Add users without class if any
            if len(users_no_class) > 0:
                no_class_data = [dict(user) for user in users_no_class]
                df_no_class = pd.DataFrame(no_class_data)
                
                # Format dates
                date_columns = ['Tanggal Daftar', 'Terakhir Hadir']
                for col in date_columns:
                    if col in df_no_class.columns:
                        df_no_class[col] = pd.to_datetime(df_no_class[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M')
                
                df_no_class.to_excel(writer, sheet_name='Tanpa Kelas', index=False)
                
                # Auto-adjust columns
                worksheet = writer.sheets['Tanpa Kelas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                total_all_students += len(users_no_class)
                active_count = len([u for u in no_class_data if u['Status'] == 'Aktif'])
                face_count = len([u for u in no_class_data if u['Face Recognition'] == 'Ya'])
                
                summary_data.append({
                    'Kelas': 'Tanpa Kelas',
                    'Jumlah Siswa': len(users_no_class),
                    'Aktif': active_count,
                    'Face Recognition': face_count
                })
            
            # Create summary sheet
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                
                # Add total row
                total_row = pd.DataFrame([{
                    'Kelas': 'TOTAL',
                    'Jumlah Siswa': df_summary['Jumlah Siswa'].sum(),
                    'Aktif': df_summary['Aktif'].sum(),
                    'Face Recognition': df_summary['Face Recognition'].sum()
                }])
                
                df_summary = pd.concat([df_summary, total_row], ignore_index=True)
                
                # Write summary as first sheet
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary sheet
                worksheet = writer.sheets['Summary']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Bold the total row
                from openpyxl.styles import Font
                last_row = len(df_summary) + 1
                for cell in worksheet[last_row]:
                    cell.font = Font(bold=True)
        
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'data_pengguna_perkelas_{timestamp}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error exporting users data: {str(e)}'
        }), 500


@app.route('/api/export/attendance/daily', methods=['GET'])
@login_required
def export_daily_attendance_excel():
    """Export daily attendance to Excel - separate sheet per class"""
    try:
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        try:
            selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }), 400
        
        conn = get_db_connection()
        
        # Get all classes
        classes = conn.execute('SELECT * FROM classes WHERE active = 1 ORDER BY name').fetchall()
        
        # Create Excel file
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = []
            
            # Process each class
            for cls in classes:
                # Get attendance for this class
                class_attendance = conn.execute('''
                    SELECT 
                        u.full_name as "Nama Lengkap",
                        a.date as "Tanggal",
                        a.time_in as "Jam Masuk",
                        a.time_out as "Jam Keluar",
                        CASE 
                            WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN 'Lengkap'
                            WHEN a.time_in IS NOT NULL AND a.time_out IS NULL THEN 'Belum Keluar'
                            ELSE 'Tidak Hadir'
                        END as "Status",
                        CASE 
                            WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN
                                PRINTF('%.1f', 
                                    CAST((julianday(a.date || ' ' || a.time_out) - 
                                          julianday(a.date || ' ' || a.time_in)) * 24 AS REAL)
                                ) || ' jam'
                            ELSE '-'
                        END as "Durasi"
                    FROM users u
                    LEFT JOIN attendance a ON u.id = a.user_id AND a.date = ?
                    WHERE u.active = 1 AND u.role != 'admin' AND u.class_id = ?
                    ORDER BY 
                        CASE WHEN a.time_in IS NOT NULL THEN 0 ELSE 1 END,
                        a.time_in ASC,
                        u.full_name ASC
                ''', (date_param, cls['id'])).fetchall()
                
                if class_attendance:
                    # Convert to DataFrame
                    df_class = pd.DataFrame([dict(row) for row in class_attendance])
                    
                    # Sanitize sheet name (max 31 chars, no special chars)
                    sheet_name = cls['name'][:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '').replace(':', '').replace('?', '')
                    
                    # Write to sheet
                    df_class.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust columns
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 25)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Calculate statistics for summary
                    total = len(df_class)
                    hadir = len(df_class[df_class['Status'] != 'Tidak Hadir'])
                    lengkap = len(df_class[df_class['Status'] == 'Lengkap'])
                    belum_keluar = len(df_class[df_class['Status'] == 'Belum Keluar'])
                    tidak_hadir = len(df_class[df_class['Status'] == 'Tidak Hadir'])
                    tingkat = round((hadir / total) * 100, 1) if total > 0 else 0
                    
                    summary_data.append({
                        'Kelas': cls['name'],
                        'Total Siswa': total,
                        'Hadir': hadir,
                        'Lengkap': lengkap,
                        'Belum Keluar': belum_keluar,
                        'Tidak Hadir': tidak_hadir,
                        'Kehadiran (%)': tingkat
                    })
            
            # Process students without class
            no_class_attendance = conn.execute('''
                SELECT 
                    u.full_name as "Nama Lengkap",
                    a.date as "Tanggal",
                    a.time_in as "Jam Masuk",
                    a.time_out as "Jam Keluar",
                    CASE 
                        WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN 'Lengkap'
                        WHEN a.time_in IS NOT NULL AND a.time_out IS NULL THEN 'Belum Keluar'
                        ELSE 'Tidak Hadir'
                    END as "Status",
                    CASE 
                        WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN
                            PRINTF('%.1f', 
                                CAST((julianday(a.date || ' ' || a.time_out) - 
                                      julianday(a.date || ' ' || a.time_in)) * 24 AS REAL)
                            ) || ' jam'
                        ELSE '-'
                    END as "Durasi"
                FROM users u
                LEFT JOIN attendance a ON u.id = a.user_id AND a.date = ?
                WHERE u.active = 1 AND u.role != 'admin' AND (u.class_id IS NULL OR u.class_id = 0)
                ORDER BY 
                    CASE WHEN a.time_in IS NOT NULL THEN 0 ELSE 1 END,
                    a.time_in ASC,
                    u.full_name ASC
            ''', (date_param,)).fetchall()
            
            if no_class_attendance:
                df_no_class = pd.DataFrame([dict(row) for row in no_class_attendance])
                df_no_class.to_excel(writer, sheet_name='Tanpa Kelas', index=False)
                
                # Auto-adjust columns
                worksheet = writer.sheets['Tanpa Kelas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 25)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add to summary
                total = len(df_no_class)
                hadir = len(df_no_class[df_no_class['Status'] != 'Tidak Hadir'])
                lengkap = len(df_no_class[df_no_class['Status'] == 'Lengkap'])
                belum_keluar = len(df_no_class[df_no_class['Status'] == 'Belum Keluar'])
                tidak_hadir = len(df_no_class[df_no_class['Status'] == 'Tidak Hadir'])
                tingkat = round((hadir / total) * 100, 1) if total > 0 else 0
                
                summary_data.append({
                    'Kelas': 'Tanpa Kelas',
                    'Total Siswa': total,
                    'Hadir': hadir,
                    'Lengkap': lengkap,
                    'Belum Keluar': belum_keluar,
                    'Tidak Hadir': tidak_hadir,
                    'Kehadiran (%)': tingkat
                })
            
            conn.close()
            
            # Create Summary sheet (first sheet)
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                
                # Add TOTAL row
                total_row = pd.DataFrame([{
                    'Kelas': 'TOTAL',
                    'Total Siswa': df_summary['Total Siswa'].sum(),
                    'Hadir': df_summary['Hadir'].sum(),
                    'Lengkap': df_summary['Lengkap'].sum(),
                    'Belum Keluar': df_summary['Belum Keluar'].sum(),
                    'Tidak Hadir': df_summary['Tidak Hadir'].sum(),
                    'Kehadiran (%)': round(
                        (df_summary['Hadir'].sum() / df_summary['Total Siswa'].sum()) * 100, 1
                    ) if df_summary['Total Siswa'].sum() > 0 else 0
                }])
                
                df_summary = pd.concat([df_summary, total_row], ignore_index=True)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary
                worksheet = writer.sheets['Summary']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Bold TOTAL row
                from openpyxl.styles import Font, PatternFill
                last_row = len(df_summary) + 1
                for cell in worksheet[last_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        output.seek(0)
        
        filename = f'kehadiran_{selected_date.strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error exporting: {str(e)}'
        }), 500        
        
@app.route('/api/export/attendance/monthly', methods=['GET'])
@login_required
def export_monthly_attendance_excel():
    """Export monthly attendance report to Excel - separate sheet per class"""
    try:
        # Get month and year parameters
        month = request.args.get('month', datetime.now().month, type=int)
        year = request.args.get('year', datetime.now().year, type=int)
        
        # Validate parameters
        if month < 1 or month > 12:
            month = datetime.now().month
        if year < 2020 or year > 2030:
            year = datetime.now().year
        
        import calendar
        
        # Get first and last day of the month
        first_day = datetime(year, month, 1).strftime('%Y-%m-%d')
        last_day = datetime(year, month, calendar.monthrange(year, month)[1]).strftime('%Y-%m-%d')
        
        conn = get_db_connection()
        
        # Get all classes
        classes = conn.execute('SELECT * FROM classes WHERE active = 1 ORDER BY name').fetchall()
        
        # Create Excel file
        output = BytesIO()
        working_days = calendar.monthrange(year, month)[1]
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = []
            all_data = []
            
            # Process each class
            for cls in classes:
                # ✅ PERBAIKAN: Tambahkan filter konsisten
                class_data = conn.execute('''
                    SELECT 
                        u.full_name as "Nama Lengkap",
                        u.username as "Username",
                        COUNT(CASE WHEN a.time_in IS NOT NULL THEN 1 END) as "Hari Hadir",
                        COUNT(CASE WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN 1 END) as "Hadir Lengkap",
                        COUNT(CASE WHEN a.time_in IS NOT NULL AND a.time_out IS NULL THEN 1 END) as "Belum Keluar",
                        COALESCE(SUM(
                            CASE WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN
                                CAST((julianday(a.date || ' ' || a.time_out) - 
                                      julianday(a.date || ' ' || a.time_in)) * 24 AS INTEGER)
                            ELSE 0 END
                        ), 0) as total_minutes,
                        MIN(a.date) as "Pertama Hadir",
                        MAX(a.date) as "Terakhir Hadir"
                    FROM users u
                    LEFT JOIN attendance a ON u.id = a.user_id 
                        AND a.date BETWEEN ? AND ?
                        AND a.time_in IS NOT NULL
                    WHERE u.active = 1 
                        AND u.role != 'admin' 
                        AND u.class_id = ?
                    GROUP BY u.id, u.full_name, u.username
                    ORDER BY u.full_name
                ''', (first_day, last_day, cls['id'])).fetchall()
                
                if class_data:
                    # Process data
                    processed_data = []
                    for row in class_data:
                        data = dict(row)
                        total_hours = round(data['total_minutes'] / 60, 1) if data['total_minutes'] else 0
                        avg_hours = round(total_hours / data['Hadir Lengkap'], 1) if data['Hadir Lengkap'] > 0 else 0
                        attendance_rate = round((data['Hari Hadir'] / working_days) * 100, 1)
                        
                        processed_row = {
                            'Nama Lengkap': data['Nama Lengkap'],
                            'Username': data['Username'],
                            'Hari Hadir': data['Hari Hadir'],
                            'Hadir Lengkap': data['Hadir Lengkap'],
                            'Belum Keluar': data['Belum Keluar'],
                            'Total Jam': total_hours,
                            'Rata-rata Jam/Hari': avg_hours,
                            'Kehadiran (%)': attendance_rate,
                            'Pertama Hadir': data['Pertama Hadir'] or '-',
                            'Terakhir Hadir': data['Terakhir Hadir'] or '-'
                        }
                        processed_data.append(processed_row)
                        all_data.append({**processed_row, 'Kelas': cls['name']})
                    
                    # Create DataFrame and write to sheet
                    df_class = pd.DataFrame(processed_data)
                    
                    # Sanitize sheet name
                    sheet_name = cls['name'][:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '').replace(':', '').replace('?', '')
                    df_class.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust columns
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 20)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add to summary
                    total = len(processed_data)
                    total_hadir = sum([d['Hari Hadir'] for d in processed_data])
                    avg_kehadiran = round((total_hadir / (total * working_days)) * 100, 1) if total > 0 else 0
                    
                    summary_data.append({
                        'Kelas': cls['name'],
                        'Jumlah Siswa': total,
                        'Total Hari Hadir': total_hadir,
                        'Rata-rata Kehadiran (%)': avg_kehadiran
                    })
            
            # Process students without class
            no_class_data = conn.execute('''
                SELECT 
                    u.full_name as "Nama Lengkap",
                    u.username as "Username",
                    COUNT(CASE WHEN a.time_in IS NOT NULL THEN 1 END) as "Hari Hadir",
                    COUNT(CASE WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN 1 END) as "Hadir Lengkap",
                    COUNT(CASE WHEN a.time_in IS NOT NULL AND a.time_out IS NULL THEN 1 END) as "Belum Keluar",
                    COALESCE(SUM(
                        CASE WHEN a.time_in IS NOT NULL AND a.time_out IS NOT NULL THEN
                            CAST((julianday(a.date || ' ' || a.time_out) - 
                                  julianday(a.date || ' ' || a.time_in)) * 24 AS INTEGER)
                        ELSE 0 END
                    ), 0) as total_minutes,
                    MIN(a.date) as "Pertama Hadir",
                    MAX(a.date) as "Terakhir Hadir"
                FROM users u
                LEFT JOIN attendance a ON u.id = a.user_id 
                    AND a.date BETWEEN ? AND ?
                    AND a.time_in IS NOT NULL
                WHERE u.active = 1 
                    AND u.role != 'admin' 
                    AND (u.class_id IS NULL OR u.class_id = 0)
                GROUP BY u.id, u.full_name, u.username
                ORDER BY u.full_name
            ''', (first_day, last_day)).fetchall()
            
            if no_class_data:
                processed_data = []
                for row in no_class_data:
                    data = dict(row)
                    total_hours = round(data['total_minutes'] / 60, 1) if data['total_minutes'] else 0
                    avg_hours = round(total_hours / data['Hadir Lengkap'], 1) if data['Hadir Lengkap'] > 0 else 0
                    attendance_rate = round((data['Hari Hadir'] / working_days) * 100, 1)
                    
                    processed_row = {
                        'Nama Lengkap': data['Nama Lengkap'],
                        'Username': data['Username'],
                        'Hari Hadir': data['Hari Hadir'],
                        'Hadir Lengkap': data['Hadir Lengkap'],
                        'Belum Keluar': data['Belum Keluar'],
                        'Total Jam': total_hours,
                        'Rata-rata Jam/Hari': avg_hours,
                        'Kehadiran (%)': attendance_rate,
                        'Pertama Hadir': data['Pertama Hadir'] or '-',
                        'Terakhir Hadir': data['Terakhir Hadir'] or '-'
                    }
                    processed_data.append(processed_row)
                    all_data.append({**processed_row, 'Kelas': 'Tanpa Kelas'})
                
                df_no_class = pd.DataFrame(processed_data)
                df_no_class.to_excel(writer, sheet_name='Tanpa Kelas', index=False)
                
                # Auto-adjust columns
                worksheet = writer.sheets['Tanpa Kelas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add to summary
                total = len(processed_data)
                total_hadir = sum([d['Hari Hadir'] for d in processed_data])
                avg_kehadiran = round((total_hadir / (total * working_days)) * 100, 1) if total > 0 else 0
                
                summary_data.append({
                    'Kelas': 'Tanpa Kelas',
                    'Jumlah Siswa': total,
                    'Total Hari Hadir': total_hadir,
                    'Rata-rata Kehadiran (%)': avg_kehadiran
                })
            
            conn.close()
            
            # Create SUMMARY sheet (first sheet)
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                
                # Add TOTAL row
                total_row = pd.DataFrame([{
                    'Kelas': 'TOTAL',
                    'Jumlah Siswa': df_summary['Jumlah Siswa'].sum(),
                    'Total Hari Hadir': df_summary['Total Hari Hadir'].sum(),
                    'Rata-rata Kehadiran (%)': round(
                        df_summary['Rata-rata Kehadiran (%)'].mean(), 1
                    )
                }])
                
                df_summary = pd.concat([df_summary, total_row], ignore_index=True)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary
                worksheet = writer.sheets['Summary']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 25)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Bold TOTAL row
                from openpyxl.styles import Font, PatternFill
                last_row = len(df_summary) + 1
                for cell in worksheet[last_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            # Create ALL DATA sheet (optional - semua user dalam 1 sheet)
            if all_data:
                df_all = pd.DataFrame(all_data)
                # Reorder columns
                cols = ['Kelas', 'Nama Lengkap', 'Username', 'Hari Hadir', 'Hadir Lengkap', 
                        'Belum Keluar', 'Total Jam', 'Rata-rata Jam/Hari', 'Kehadiran (%)', 
                        'Pertama Hadir', 'Terakhir Hadir']
                df_all = df_all[cols]
                df_all.to_excel(writer, sheet_name='Semua Data', index=False)
                
                # Auto-adjust columns
                worksheet = writer.sheets['Semua Data']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        filename = f'laporan_bulanan_{calendar.month_name[month]}_{year}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error exporting monthly report: {str(e)}'
        }), 500
        
        
@app.route('/api/cleanup/photos', methods=['POST'])
@login_required
def api_cleanup_photos():
    """Manual cleanup of old attendance photos (admin only)"""
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Access denied. Admin only.'}), 403
    
    try:
        deleted_count = cleanup_old_attendance_photos(days_to_keep=7)
        
        return jsonify({
            'success': True,
            'message': f'Berhasil menghapus {deleted_count} foto lama (>7 hari)',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500

@app.route('/api/cleanup/stats', methods=['GET'])
@login_required
def api_cleanup_stats():
    """Get cleanup statistics (admin only)"""
    if session.get('username') != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        conn = get_db_connection()
        cutoff_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        
        # Count old photos
        old_photos = conn.execute('''
            SELECT 
                COUNT(CASE WHEN photo_path IS NOT NULL THEN 1 END) as check_in_photos,
                COUNT(CASE WHEN photo_path_out IS NOT NULL THEN 1 END) as check_out_photos
            FROM attendance 
            WHERE date < ?
        ''', (cutoff_date,)).fetchone()
        
        total_old = old_photos['check_in_photos'] + old_photos['check_out_photos']
        
        # Count recent photos
        recent_photos = conn.execute('''
            SELECT 
                COUNT(CASE WHEN photo_path IS NOT NULL THEN 1 END) as check_in_photos,
                COUNT(CASE WHEN photo_path_out IS NOT NULL THEN 1 END) as check_out_photos
            FROM attendance 
            WHERE date >= ?
        ''', (cutoff_date,)).fetchone()
        
        total_recent = recent_photos['check_in_photos'] + recent_photos['check_out_photos']
        
        conn.close()
        
        return jsonify({
            'success': True,
            'stats': {
                'old_photos_count': total_old,
                'recent_photos_count': total_recent,
                'cutoff_date': cutoff_date,
                'retention_days': 7
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500
        
        
# Tambahkan endpoint ini ke app.py Anda (letakkan di bagian API routes)

@app.route('/api/classes/list', methods=['GET'])
@login_required
def api_classes_list():
    """API to get list of classes"""
    try:
        conn = get_db_connection()
        
        classes = conn.execute('''
            SELECT id, name, description, active, created_at
            FROM classes 
            WHERE active = 1
            ORDER BY name ASC
        ''').fetchall()
        
        classes_list = []
        for cls in classes:
            classes_list.append({
                'id': cls['id'],
                'name': cls['name'],
                'description': cls['description'] if cls['description'] else '',
                'active': cls['active'],
                'created_at': cls['created_at']
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'classes': classes_list,
            'count': len(classes_list)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'classes': []
        }), 500

@app.route('/')
def index():
    """Main dashboard page"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Redirect admin to manage user page
    if session.get('username') == 'admin':
        return redirect(url_for('users_dashboard'))
    
    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    
    # Get today's attendance
    today = datetime.now().strftime("%Y-%m-%d")
    attendance = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? AND date = ?',
        (session['user_id'], today)
    ).fetchone()
    
    # Get attendance statistics
    stats = conn.execute(
        '''SELECT 
           COUNT(*) as total_days,
           SUM(CASE WHEN time_in IS NOT NULL THEN 1 ELSE 0 END) as present_days
           FROM attendance WHERE user_id = ?''',
        (session['user_id'],)
    ).fetchone()
    
    # Check if user has face recognition enabled
    face_enabled = conn.execute(
        'SELECT COUNT(*) FROM face_data WHERE user_id = ? AND active = 1',
        (session['user_id'],)
    ).fetchone()[0] > 0
    
    # Check if should show face setup reminder
    show_face_reminder = session.pop('show_face_reminder_on_dashboard', False) and not face_enabled
    
    conn.close()
    
    return render_template('index.html',    
                         user=user, 
                         attendance=attendance, 
                         stats=stats,
                         face_enabled=face_enabled,
                         show_face_reminder=show_face_reminder,
                         face_recognition_available=FACE_RECOGNITION_AVAILABLE)


if __name__ == '__main__':
    import os
    
    # Create directories if they don't exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['FACES_FOLDER'], exist_ok=True)
    
    # Auto cleanup old attendance photos (7 days retention)
    print("🧹 Running photo cleanup...")
    deleted = cleanup_old_attendance_photos(days_to_keep=7)
    if deleted > 0:
        print(f"✅ Deleted {deleted} old photos (>7 days)")
    else:
        print("✅ No old photos to delete")
    
    # Initialize web registration
    if FACE_RECOGNITION_AVAILABLE:
        init_web_registration(app)
        print("✓ Face recognition enabled")
    else:
        print("⚠ Face recognition disabled - install required packages")
    
    use_ssl = os.path.exists('cert.pem') and os.path.exists('key.pem')
    
    if use_ssl:
        import ssl
        context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
        context.load_cert_chain('cert.pem', 'key.pem')
        print("✓ Running with HTTPS")
        print("Access via: https://172.15.3.16:5000 or https://localhost:5000")
        app.run(debug=True, host='0.0.0.0', port=5000, ssl_context=context)
    else:
        print("⚠ Running with HTTP (no SSL certificate found)")
        print("⚠ Geolocation may not work on IP address access")
        print("Access via: http://localhost:5000 (recommended)")
        print("Or generate SSL certificate with:")
        print("  openssl req -x509 -newkey rsa:4096 -nodes -out cert.pem -keyout key.pem -days 365")
        app.run(debug=True, host='0.0.0.0', port=5000)